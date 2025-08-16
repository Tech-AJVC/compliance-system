"""
Excel Generation Service for Unit Allotment Sheets
"""
import os
import logging
from typing import List, Dict, Any
from datetime import datetime, date
from pathlib import Path
import tempfile
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from ..models.unit_allotment import UnitAllotment
from ..utils.s3_storage import S3DocumentStorage

logger = logging.getLogger(__name__)

class UnitAllotmentExcelGenerator:
    """
    Service for generating Excel sheets for unit allotments with regulatory compliance formatting.
    """
    
    def __init__(self):
        self.s3_storage = S3DocumentStorage()
        
        # Excel formatting constants
        self.HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.DATA_FONT = Font(name='Arial', size=10)
        self.BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
        self.RIGHT_ALIGNMENT = Alignment(horizontal='right', vertical='center')
    
    def generate_allotment_sheet(self, 
                                 allotments: List[UnitAllotment],
                                 fund_name: str,
                                 drawdown_quarter: str) -> str:
        """
        Generate Excel sheet for unit allotments with regulatory formatting.
        
        Args:
            allotments: List of UnitAllotment objects
            fund_name: Name of the fund
            drawdown_quarter: Quarter for the drawdown (e.g., "FY25Q1")
            
        Returns:
            str: S3 URL of the generated Excel file
        """
        if not allotments:
            raise ValueError("No allotments provided for Excel generation")
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Unit Allotment"
        
        # Set up headers and data
        self._setup_headers(worksheet)
        self._populate_data(worksheet, allotments)
        self._apply_formatting(worksheet, len(allotments))
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(
            suffix='.xlsx', 
            delete=False, 
            prefix=f'unit_allotment_{drawdown_quarter}_'
        ) as temp_file:
            workbook.save(temp_file.name)
            temp_file_path = temp_file.name
        
        try:
            # Upload to S3
            s3_key = self._generate_s3_key(fund_name, drawdown_quarter)
            upload_result = self.s3_storage.upload_file(
                local_file_path=temp_file_path,
                s3_key=s3_key,
                metadata={
                    'document_type': 'unit_allotment',
                    'fund_name': fund_name,
                    'drawdown_quarter': drawdown_quarter,
                    'total_lps': str(len(allotments)),
                    'generated_date': datetime.utcnow().isoformat()
                },
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            logger.info(f"Unit allotment Excel generated and uploaded to S3: {s3_key}")
            return upload_result['s3_url']
            
        finally:
            # Clean up temporary file
            try:
                os.unlink(temp_file_path)
            except OSError:
                logger.warning(f"Could not delete temporary file: {temp_file_path}")
    
    def _setup_headers(self, worksheet):
        """Set up Excel headers based on the exact order from the provided template."""
        headers = [
            'FIRST-HOLDER-NAME', 'FIRST-HOLDER-PAN', 'SECOND-HOLDER-NAME', 'SECOND-HOLDER-PAN',
            'THIRD-HOLDER-NAME', 'THIRD-HOLDER-PAN', 'Depository', 'DPID', 'CLID',
            'ALLOTED UNIT', 'NAV/F. VALUE', 'DATE OF ALLOTMENT', 'COMMITTED AMT',
            'DRAWDOWN AMOUNT', 'MGMT FEES', 'AMT ACCEPTED', 'BANK ACCOUNT NO', 'BANK NAME',
            'MICR CODE', 'IFSC CODE', 'STAMP DUTY', 'STATUS', 'DRAWDOWN DATE', 
            'DRAWDOWN QUARTER', 'NOTIFICATION', 'FLAG'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER_ALIGNMENT
            cell.border = self.BORDER
    
    def _populate_data(self, worksheet, allotments: List[UnitAllotment]):
        """Populate worksheet with allotment data in the exact order from the template."""
        for row_num, allotment in enumerate(allotments, 2):  # Start from row 2
            data_row = [
                allotment.first_holder_name,  # FIRST-HOLDER-NAME
                allotment.first_holder_pan or '',  # FIRST-HOLDER-PAN
                allotment.second_holder_name or '',  # SECOND-HOLDER-NAME
                allotment.second_holder_pan or '',  # SECOND-HOLDER-PAN
                allotment.third_holder_name or '',  # THIRD-HOLDER-NAME
                allotment.third_holder_pan or '',  # THIRD-HOLDER-PAN
                allotment.depository or '',  # Depository
                allotment.dpid or '',  # DPID
                allotment.clid or '',  # CLID
                allotment.allotted_units,  # ALLOTED UNIT
                allotment.nav_value,  # NAV/F. VALUE (now integer)
                allotment.date_of_allotment.strftime('%d/%m/%Y') if allotment.date_of_allotment else '',  # DATE OF ALLOTMENT
                float(allotment.committed_amt),  # COMMITTED AMT
                float(allotment.drawdown_amount),  # DRAWDOWN AMOUNT
                float(allotment.mgmt_fees),  # MGMT FEES
                float(allotment.amt_accepted),  # AMT ACCEPTED
                allotment.bank_account_no or '',  # BANK ACCOUNT NO
                allotment.bank_account_name or '',  # BANK NAME
                allotment.micr_code or '',  # MICR CODE
                allotment.bank_ifsc or '',  # IFSC CODE
                float(allotment.stamp_duty),  # STAMP DUTY
                allotment.status,  # STATUS
                allotment.drawdown_date.strftime('%d/%m/%Y') if allotment.drawdown_date else '',  # DRAWDOWN DATE
                allotment.drawdown_quarter,  # DRAWDOWN QUARTER
                '',  # NOTIFICATION (empty)
                ''   # FLAG (empty)
            ]
            
            for col_num, value in enumerate(data_row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = self.DATA_FONT
                cell.border = self.BORDER
                
                # Apply specific formatting based on data type and new column positions
                if isinstance(value, float) and col_num in [13, 14, 15, 16, 21]:  # Financial columns: COMMITTED AMT, DRAWDOWN AMOUNT, MGMT FEES, AMT ACCEPTED, STAMP DUTY
                    cell.number_format = '#,##0.00'
                    cell.alignment = self.RIGHT_ALIGNMENT
                elif col_num in [10, 11]:  # Unit columns (integers): ALLOTED UNIT, NAV/F. VALUE
                    cell.number_format = '#,##0'
                    cell.alignment = self.RIGHT_ALIGNMENT
                else:
                    cell.alignment = self.CENTER_ALIGNMENT
    
    def _apply_formatting(self, worksheet, data_rows: int):
        """Apply final formatting to the worksheet."""
        # Auto-adjust column widths with better handling for long column names
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Enhanced width calculation for better readability
            if max_length > 10:  # For longer column names, add more buffer
                buffer = 4
            else:
                buffer = 2
            
            # Set minimum width based on content, with better minimums for readability
            base_width = max_length + buffer
            min_width = 15 if max_length > 13 else 12  # Higher minimum for long headers
            max_width = 60  # Increased max width for very long content
            
            adjusted_width = min(max(base_width, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
        
        # Add data validation and protection if needed
        # This could be extended for additional regulatory requirements
    
    def _generate_s3_key(self, fund_name: str, drawdown_quarter: str) -> str:
        """Generate S3 key for the allotment sheet."""
        # Clean fund name for file path
        safe_fund_name = "".join(c for c in fund_name if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_fund_name = safe_fund_name.replace(' ', '_')
        
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f"unit_allotment_{safe_fund_name}_{drawdown_quarter}_{timestamp}.xlsx"
        
        return f"{safe_fund_name}/{drawdown_quarter}/Unit allotment/{filename}"
    
    def generate_individual_certificate(self, allotment: UnitAllotment, fund_name: str) -> str:
        """
        Generate individual unit certificate for an LP.
        
        Args:
            allotment: UnitAllotment object for a single LP
            fund_name: Name of the fund
            
        Returns:
            str: S3 URL of the generated certificate
        """
        # This method could be expanded to generate individual certificates
        # For now, we'll create a simple single-LP Excel sheet
        return self.generate_allotment_sheet([allotment], fund_name, allotment.drawdown_quarter)
    
    def validate_allotment_data(self, allotments: List[UnitAllotment]) -> Dict[str, List[str]]:
        """
        Validate allotment data before Excel generation.
        
        Args:
            allotments: List of UnitAllotment objects to validate
            
        Returns:
            Dict[str, List[str]]: Dictionary of validation errors by category
        """
        errors = {
            'missing_required_fields': [],
            'invalid_calculations': [],
            'data_inconsistencies': []
        }
        
        for i, allotment in enumerate(allotments):
            lp_identifier = f"LP {i+1} ({allotment.first_holder_name})"
            
            # Check required fields
            if not allotment.first_holder_name:
                errors['missing_required_fields'].append(f"{lp_identifier}: Missing LP name")
            
            if not allotment.drawdown_amount or allotment.drawdown_amount <= 0:
                errors['missing_required_fields'].append(f"{lp_identifier}: Invalid drawdown amount")
            
            if not allotment.allotted_units or allotment.allotted_units <= 0:
                errors['invalid_calculations'].append(f"{lp_identifier}: Invalid unit allocation")
            
            # Validate calculations
            if allotment.nav_value and allotment.drawdown_amount:
                expected_units = int(allotment.drawdown_amount / allotment.nav_value)
                if abs(allotment.allotted_units - expected_units) > 1:  # Allow for rounding differences
                    errors['invalid_calculations'].append(
                        f"{lp_identifier}: Unit calculation mismatch (expected ~{expected_units}, got {allotment.allotted_units})"
                    )
        
        # Remove empty error categories
        return {k: v for k, v in errors.items() if v}
    
    def get_allotment_summary(self, allotments: List[UnitAllotment]) -> Dict[str, Any]:
        """
        Generate summary statistics for the allotment.
        
        Args:
            allotments: List of UnitAllotment objects
            
        Returns:
            Dict[str, Any]: Summary statistics
        """
        if not allotments:
            return {}
        
        total_units = sum(a.allotted_units for a in allotments)
        total_amount = sum(a.drawdown_amount for a in allotments)
        total_mgmt_fees = sum(a.mgmt_fees for a in allotments)
        total_stamp_duty = sum(a.stamp_duty for a in allotments)
        
        return {
            'total_lps': len(allotments),
            'total_units_allocated': total_units,
            'total_amount_allocated': float(total_amount),
            'total_management_fees': float(total_mgmt_fees),
            'total_stamp_duty': float(total_stamp_duty),
            'drawdown_quarter': allotments[0].drawdown_quarter,
            'generation_date': datetime.utcnow().isoformat()
        }