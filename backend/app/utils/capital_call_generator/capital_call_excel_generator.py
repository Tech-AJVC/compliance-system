"""
Capital Call Notice Generator - Excel Version using openpyxl
"""
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# from openpyxl.drawing import image  # Commented out for now
from datetime import datetime
import logging
from typing import Dict, Any

logger = logging.getLogger(__name__)

class CapitalCallExcelGenerator:
    def __init__(self, output_dir: str = "uploads/capital_calls"):
        """
        Initialize the Excel generator
        
        Args:
            output_dir: Directory to save generated Excel files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_capital_call_excel(self, data: Dict[str, Any], filename: str = None) -> str:
        """
        Generate capital call notice as Excel file
        
        Args:
            data: Dictionary containing all the dynamic data for the template
            filename: Optional custom filename
            
        Returns:
            Path to the generated Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"capital_call_{data.get('investor', 'unknown').replace(' ', '_')}_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Capital Call"
        
        # Define styles
        title_font = Font(name='Arial', size=24, bold=True, color='FF0000')
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=11)
        
        # Add AJVC logo placeholder (you can add actual logo later)
        ws.merge_cells('A2:C3')
        ws['A2'] = "AJVC"
        ws['A2'].font = Font(name='Arial', size=16, bold=True, color='FF0000')
        
        # Title
        ws['A5'] = "Capital Call"
        ws['A5'].font = title_font
        ws.merge_cells('A5:F5')
        ws['A5'].alignment = Alignment(horizontal='left')
        
        # Notice Date and Investor
        ws['A7'] = "Notice date"
        ws['A7'].font = header_font
        ws['A8'] = data.get('notice_date', '')
        ws['A8'].font = normal_font
        
        ws['D7'] = "Investor"
        ws['D7'].font = header_font
        ws['D8'] = data.get('investor', '')
        ws['D8'].font = normal_font
        
        # Personal message
        current_row = 10
        messages = [
            f"Dear {data.get('investor', '')},",
            "",
            "Thank you for being AJVC's first supporters as we begin this journey. We are grateful for",
            "your support from Day 1.",
            "",
            "We are writing to notify you, as per the Contribution Agreement entered with AJVC Fund.",
            "",
            f"The Fund is calling capital from you in the amount of INR {data.get('amount_due', 0):,} - which is",
            f"{(data.get('amount_due', 0)/data.get('total_commitment', 1)*100):.0f}% of your committed capital.",
            "",
            f"Contribution is due on or before {data.get('contribution_due_date', '')} and must be wired in immediately.",
            "",
            "The wire instructions are as follows:"
        ]
        
        for message in messages:
            if message:
                ws[f'A{current_row}'] = message
                ws[f'A{current_row}'].font = normal_font
            current_row += 1
        
        # Bank details
        bank_details = [
            ("Bank Name:", data.get('bank_name', '')),
            ("IFSC:", data.get('ifsc', '')),
            ("Acct Name:", data.get('acct_name', '')),
            ("Acct Number:", data.get('acct_number', '')),
            ("Bank Contact:", data.get('bank_contact', '')),
            ("Phone:", data.get('phone', ''))
        ]
        
        for label, value in bank_details:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].font = header_font
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].font = normal_font
            current_row += 1
        
        current_row += 1  # Add space
        
        # Commitment summary table
        ws[f'A{current_row}'] = "Investor Commitment Summary"
        ws[f'A{current_row}'].font = header_font
        ws[f'E{current_row}'] = "Amount"
        ws[f'E{current_row}'].font = header_font
        current_row += 1
        
        # Table data
        table_data = [
            ("Total commitment", f"₹ {data.get('total_commitment', 0):,}"),
            ("Amount called-up and paid till date", f"₹ {data.get('amount_called_up', 0):,}"),
            ("Remaining commitment after this drawdown", f"₹ {data.get('remaining_commitment', 0):,}"),
            ("Amount due", f"₹ {data.get('amount_due', 0):,}")
        ]
        
        # Add borders to table
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        table_start_row = current_row - 1
        for desc, amount in table_data:
            ws[f'A{current_row}'] = desc
            ws[f'A{current_row}'].font = normal_font
            ws[f'A{current_row}'].border = thin_border
            ws[f'E{current_row}'] = amount
            ws[f'E{current_row}'].font = normal_font
            ws[f'E{current_row}'].border = thin_border
            current_row += 1
        
        # Add border to header row
        ws[f'A{table_start_row}'].border = thin_border
        ws[f'E{table_start_row}'].border = thin_border
        
        current_row += 1  # Add space
        
        # Final instructions
        final_messages = [
            f"Please instruct the financial institution ({data.get('bank_name', '')}) handling the wire",
            "transfer to include your name, as a Limited Partner of the Fund.",
            "",
            "We plan to do quarterly drawdowns at the beginning of each quarter.",
            f"Forecasted next quarter ({data.get('forecast_next_quarter_period', '')}) drawdown: {data.get('forecast_next_quarter', 0):.0f}% of the committed amount."
        ]
        
        for message in final_messages:
            if message:
                ws[f'A{current_row}'] = message
                ws[f'A{current_row}'].font = normal_font
            current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file
        file_path = self.output_dir / filename
        wb.save(file_path)
        
        logger.info(f"Generated Excel capital call notice: {file_path}")
        return str(file_path)

def generate_capital_call_excel(data: Dict[str, Any], output_path: str = None) -> str:
    """
    Convenience function to generate capital call Excel
    
    Args:
        data: Dictionary containing all the dynamic data
        output_path: Optional output path
        
    Returns:
        Path to generated file
    """
    generator = CapitalCallExcelGenerator()
    
    if output_path:
        filename = Path(output_path).name
    else:
        filename = None
        
    return generator.generate_capital_call_excel(data, filename)